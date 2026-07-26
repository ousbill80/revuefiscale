"""Lecteur balance Excel (.xlsx)."""
from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from backend.socle.erreurs import ErreurLectureBalance
from backend.socle.modeles import LigneBalance


def _dec(valeur: object, champ: str, ligne_no: int) -> Decimal:
    if valeur is None or valeur == "":
        return Decimal("0")
    if isinstance(valeur, (int, float, Decimal)):
        return Decimal(str(valeur))
    texte = str(valeur).strip().replace(" ", "").replace(",", ".")
    if texte == "":
        return Decimal("0")
    try:
        return Decimal(texte)
    except InvalidOperation as e:
        raise ErreurLectureBalance(
            f"ligne {ligne_no} : {champ} non numerique ({valeur!r})"
        ) from e


def parser_balance_xlsx(contenu: bytes) -> list[LigneBalance]:
    """Parse un classeur xlsx : colonnes compte, libelle, debit, credit.

    Premiere feuille. Entete optionnel (ligne contenant 'compte').
    """
    if not contenu:
        raise ErreurLectureBalance("fichier Excel vide")
    try:
        wb = load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
    except Exception as e:
        raise ErreurLectureBalance(f"Excel illisible : {e}") from e
    ws = wb.active
    if ws is None:
        raise ErreurLectureBalance("aucune feuille Excel")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ErreurLectureBalance("feuille Excel vide")

    debut = 0
    i_compte, i_libelle, i_debit, i_credit = 0, 1, 2, 3
    premiere = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    if "compte" in premiere:
        debut = 1
        try:
            i_compte = premiere.index("compte")
            i_libelle = premiere.index("libelle") if "libelle" in premiere else -1
            i_debit = premiere.index("debit")
            i_credit = premiere.index("credit")
        except ValueError as e:
            raise ErreurLectureBalance(
                "entete Excel attendu : compte,libelle,debit,credit"
            ) from e

    resultat: list[LigneBalance] = []
    for no, raw in enumerate(rows[debut:], start=debut + 1):
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        if len(raw) <= max(i_compte, i_debit, i_credit):
            raise ErreurLectureBalance(f"ligne {no} : colonnes insuffisantes")
        compte_val = raw[i_compte]
        if compte_val is None or str(compte_val).strip() == "":
            raise ErreurLectureBalance(f"ligne {no} : compte vide")
        # Excel peut lire un compte comme float (401.0)
        compte = str(compte_val).strip()
        if compte.endswith(".0") and compte.replace(".", "", 1).isdigit():
            compte = compte[:-2]
        libelle = None
        if i_libelle >= 0 and i_libelle < len(raw) and raw[i_libelle] is not None:
            libelle = str(raw[i_libelle]).strip() or None
        debit = _dec(raw[i_debit] if i_debit < len(raw) else 0, "debit", no)
        credit = _dec(raw[i_credit] if i_credit < len(raw) else 0, "credit", no)
        resultat.append(
            LigneBalance(compte=compte, libelle=libelle, debit=debit, credit=credit)
        )

    if not resultat:
        raise ErreurLectureBalance("aucune ligne de compte")
    return resultat
